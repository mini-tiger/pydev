import requests
import json, time, decimal, os
from models import data
from text2vec import Similarity, SimilarityType
from tools import tool
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from openpyxl.comments import *
from rouge import Rouge

sim_model = Similarity("shibing624/text2vec-base-chinese", SimilarityType.COSINE)


class TestReqScore():
    def __init__(self, **kwargs):
        self.src_json_file = kwargs["src_struct"]["file"]
        self.data_type = kwargs["src_struct"]["data_type"]
        self.prompt_prefix = kwargs["src_struct"]["prompt_prefix"]
        self.max_use = kwargs["src_struct"]["max_use"]
        self.url = "http://120.133.83.145:8001/v1/chat/completions"
        self.payload_list = []
        self.headers = {
            'Content-Type': 'application/json'
        }
        self.data = []
        self.result_list = []

        self.out_csv = "/mnt/AI_Json/%s.csv" % os.path.basename(self.src_json_file)
        self.out_xlsx = "/mnt/AI_Json/%s.xlsx" % os.path.basename(self.src_json_file)
        self.rouge = Rouge()

    def compute(self, answer_std, answer_real):
        score = sim_model.get_score(answer_std, answer_real)
        dec = decimal.Decimal(score)
        score = dec.quantize(decimal.Decimal('0.00'), rounding=decimal.ROUND_DOWN)

        # 返回答案中找到 真实答案 返回1，返回答案中包括 抱歉 无法 返回0
        if answer_real.find(answer_std) != -1:
            # print(f"'{answer_real}' 包含子字符串 '{answer_std}'")
            return decimal.Decimal(1)

        if self.data_type != "机器翻译" and (
                answer_real.find("抱歉") != -1 or answer_real.find("无法") != -1 or answer_real.find("对不起") != -1):
            return 0

            # print(f"'{answer_real}' 不包含子字符串 '{answer_std}'")
        return score

    def rouge_compute(self, answer_std, answer_real):
        rouge_scores = self.rouge.get_scores(hyps=answer_std, refs=answer_real, avg=False)

        if len(rouge_scores) < 1:
            return 0
        if "rouge-l" in rouge_scores[0]:
            # print(rouge_scores[0])
            dec = decimal.Decimal(rouge_scores[0]["rouge-l"]["f"])
            return dec.quantize(decimal.Decimal('0.00'), rounding=decimal.ROUND_DOWN)

        else:
            return 0

    def format_json(self):
        with open(self.src_json_file, "r", encoding="utf-8") as json_file:
            self.data = json.load(json_file)

    def request_model(self):
        max_use = self.max_use if self.max_use > 0 else len(self.data)

        for d in self.data[0:max_use]:
            p = d["prompt"] if len(d["prompt"]) > 3 else d["question_std"]
            prompt = "%s%s" % (self.prompt_prefix, p)
            # payload = json.dumps({
            #     "prompt": prompt,
            #     "history": []
            # })
            payload = json.dumps({
                "model": "baichuan-13b-chat",
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }]
            })

            singe_qa = data.PromptQA()
            singe_qa.type = self.data_type
            singe_qa.put_prompt(prompt=prompt)
            try:
                response = requests.request("POST", self.url, headers=self.headers, data=payload, timeout=10)
                res = json.loads(response.text).get('choices')[0].get("message").get("content")
                # print(res)
                # 检查响应内容中是否包含特定字段

                if res is not None:
                    # print(f"响应包含字段")

                    singe_qa.put_data(valid=True, answers_std=d["answers_std"], answers_real=res,
                                      question_std=d["question_std"])
                    singe_qa.err = ""
                    score_vec = self.compute(d["answers_std"], res)
                    score_rouge = self.rouge_compute(d["answers_std"], res)
                    print("score_vec:%s ,rouge-l:%s,prompt: %s,answer_std: %s,answer_real:%s" % (score_vec, score_rouge,
                                                                                                 prompt,
                                                                                                 singe_qa.answers_std,
                                                                                                 singe_qa.answers_real))
                    singe_qa.vec_score = score_vec
                    singe_qa.rouge_score = score_rouge

                else:
                    print(f"url响应不包含字段")
                    singe_qa.put_data(valid=False, answers_std=d["answers_std"], answers_real=res,
                                      question_std=d["question_std"])
                    singe_qa.err = f"响应不包含字段"
            except Exception as e:
                print("exception:%s" % e)
                singe_qa.put_data(valid=False, answers_std=d["answers_std"], answers_real="",
                                  question_std=d["question_std"])
                singe_qa.err = e.__str__()
            self.result_list.append(singe_qa)

    def process(self):
        self.format_json()
        self.request_model()
        # self.write_csv()
        # self.write_xlsx()

    def write_xlsx(self):
        tool.write_xlsx(file=self.out_xlsx, data_list=self.result_list,
                        fieldnames=["type", "vec_score", "rouge_score", "question_std", "answers_std", "answers_real",
                                    "valid", "prompt",
                                    "err"])

    def write_csv(self):
        tool.write_csv(file=self.out_csv, data_list=self.result_list,
                       fieldnames=["type", "vec_score", "rouge_score", "question_std", "answers_std", "answers_real",
                                   "valid", "prompt",
                                   "err"])

    def get_res_data(self):
        return self.result_list


class BatchProcess():
    def __init__(self, *args, **kwargs):
        self.file_list = kwargs["file_list"]
        # self.fieldnames = ["type",  "vec_score","rouge_score", "question_std", "answers_std", "answers_real", "valid", "prompt", "err"]
        self.all_data = []
        self.out_file = os.path.join(kwargs["out_dir"], "Ares.xlsx")
        self.report_avg = []
        self.new_sheetname = tool.DateStr()
        # print(self.out_file)

    def compute_avg(self, data):
        avg = 0
        dtype = ""
        for i in data:
            avg += i.vec_score
            dtype = i.type
        self.report_avg.append([self.new_sheetname, dtype, "%d" % (avg / len(data))])

    def process(self):
        for f in self.file_list:
            start = time.time()
            ss = TestReqScore(src_struct=f)
            ss.process()
            currentData = ss.get_res_data()
            self.all_data.extend(currentData)
            self.compute_avg(currentData)
            end = time.time()
            print("完成时间: %f s" % (end - start))  #
        # print(self.all_data)
        # print(len(self.all_data))
        self.write_wrap_xlsx()

    def write_wrap_xlsx(self):
        self.write_xlsx(field_dict={"type": {"comment": "问题类型", "w": 18},
                                    "vec_score": {"comment": "余弦匹配分数", "w": 18},
                                    "rouge_score": {"comment": "rouge分数", "w": 18},
                                    "question_std": {"comment": "标准问题", "w": 80},
                                    "answers_std": {"comment": "标准答案", "w": 80},
                                    "answers_real": {"comment": "实际模型返回答案", "w": 80},
                                    "valid": {"comment": "有效执行", "w": 10},
                                    "prompt": {"comment": "请求API参数,封装标准问题", "w": 80},
                                    "err": {"comment": "执行中报错", "w": 15}
                                    }
                        )

    def write_xlsx(self, field_dict):
        fieldnames = list(field_dict.keys())
        # Specify XLSX file path
        xlsx_file_path = self.out_file
        workbook = Workbook()
        fontObj = Font(name=u'微软雅黑', bold=True, size=14)
        if os.path.exists(xlsx_file_path) is False:
            # Get the active worksheet
            sheet = workbook.active
            sheet.title = self.new_sheetname
        # Write column names
        else:
            workbook = load_workbook(xlsx_file_path)
            # 添加一个新的sheet页，可以通过指定标题来命名它
            sheet = workbook.create_sheet(title=tool.DateStr())
            workbook.active = sheet

        if 'Report' in workbook.sheetnames:
            reportsheet = workbook['Report']
            # print(reportsheet.max_row)

            # ws = workbook["统计"]
            # pivot = ws._pivots[0]  # any will do as they share the same cache
            # pivot.cache.refreshOnLoad = True
        else:
            reportsheet = workbook.create_sheet('Report', 0)
            reportsheet.append(["日期", " 类型", "vec平均分"])
            reportsheet.column_dimensions[get_column_letter(1)].width = 20
            reportsheet.column_dimensions[get_column_letter(2)].width = 20
            reportsheet.column_dimensions[get_column_letter(3)].width = 20
            sheet['A1'].font = fontObj
            sheet['B1'].font = fontObj
            sheet['C1'].font = fontObj
        # reportsheet=workbook.active
        for d in self.report_avg:
            reportsheet.append(d)

        sheet.append(fieldnames)
        sheet.freeze_panes = 'A2'

        format_res = []
        for i in self.all_data:
            row = []
            for field in fieldnames:
                # print(field,getattr(i,field))
                row.append(getattr(i, field))
            format_res.append(row)

        # Write data from PromptQA objects to the worksheet
        for row in format_res:
            sheet.append(row)

        # 将整个表的行高设置为 50，列宽设置为 30；
        # 设置连续行行高：
        for r in range(2, len(self.all_data) + 2):  # 注意，行和列的序数，都是从1开始
            sheet.row_dimensions[r].height = 80  #
        # 设置连续列列宽：
        for c in range(1, len(fieldnames) + 1):  # 注意，列序数从1开始，但必须转变为A\B等字母
            w = get_column_letter(c)  # 把列序数转变为字母
            sheet.column_dimensions[w].width = field_dict[fieldnames[c - 1]]["w"]
            sheet[w + '1'].font = fontObj
            sheet[w + '1'].comment = comments.Comment(field_dict[fieldnames[c - 1]]["comment"], "")

        # 垂直居中
        # nrows = sheet.max_row  # 获得行数
        # ncols = sheet.max_column
        # for i in range(nrows):
        #     for j in range(ncols):
        #         sheet.cell(row=i + 1, column=j + 1).alignment = Alignment( vertical='center')

        currentCell = sheet.cell(row=1, column=1)  # or currentCell = ws['A1']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        # 边框
        tool.format_border(sheet, "A", "1", get_column_letter(len(fieldnames)), len(self.all_data) + 1)
        ##设置全部单元 自动换行
        for key in list(sheet._cells.keys()): sheet._cells[key].alignment = Alignment(wrapText=True)

        # sheet 排序
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save the workbook to XLSX file
        workbook.save(xlsx_file_path)

        print("Data written to", xlsx_file_path)


if __name__ == "__main__":
    current_directory = os.path.dirname(__file__)
    filedir = os.path.join(current_directory, "format_json", "json_data")
    file_list = [
        {"data_type": "艺术", "file": os.path.join(filedir, "cmmlu", "arts.csv.json"), "prompt_prefix": "",
         "max_use": 10},
        # {"data_type": "医学", "file": os.path.join(filedir, "cmmlu", "anatomy.csv.json"), "prompt_prefix": "",
        #  "max_use": 10},
        # {"data_type": "法律", "file": os.path.join(filedir, "ceval", "law_val.csv.json"), "prompt_prefix": "",
        #  "max_use": 10},
        # {"data_type": "艺术", "file": os.path.join(filedir, "ceval", "art_studies_val.csv.json"), "prompt_prefix": "",
        #  "max_use": 10},
        # {"data_type": "财务", "file": os.path.join(filedir, "ceval", "accountant_val.csv.json"), "prompt_prefix": "",
        #  "max_use": 10},
        # {"data_type": "科学常识", "file": os.path.join(filedir, "科学.json"), "prompt_prefix": "", "max_use": 0},
        # # {"data_type": "数据中心标注", "file": os.path.join(filedir, "Colo_20230830_modify.xlsx.json"), "prompt_prefix": "",
        # #  "max_use": 10},
        # {"data_type": "数据中心", "file": os.path.join(filedir, "datacenter.json"), "prompt_prefix": "", "max_use": 10},
        # {"data_type": "数据提取", "file": os.path.join(filedir, "chouqu.json"), "prompt_prefix": "", "max_use": 10},
        # {"data_type": "数学运算", "file": os.path.join(filedir, "math.json"), "prompt_prefix": "不需要计算过程,只返回答案字母: 计算 ",
        #  "max_use": 10},
        {"data_type": "机器翻译", "file": os.path.join(filedir, "translation.json"), "prompt_prefix": "翻译英文至中文 ",
         "max_use": 10}
    ]

    # file_list = ["超互联", "荷丹", ]
    BatchProcess(file_list=file_list, out_dir="/mnt/AI_Json").process()
