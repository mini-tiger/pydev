import os.path

import pandas as pd
from datetime import datetime
import json, csv
import numpy as np
from text2vec import Similarity, SimilarityType
import decimal, sys

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# Check current encoding
current_encoding = sys.getdefaultencoding()
print("Current encoding:", current_encoding)

from openpyxl.styles import Side, Border, colors


# 定义边框样式
def my_border(t_border, b_border, l_border, r_border):
    border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                    bottom=Side(border_style=b_border, color=colors.BLACK),
                    left=Side(border_style=l_border, color=colors.BLACK),
                    right=Side(border_style=r_border, color=colors.BLACK))
    return border


# 初始化制定区域边框为所有框线
def format_border(sheet, s_column, s_index, e_column, e_index):
    for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
        for cell in row:
            cell.border = my_border('thin', 'thin', 'thin', 'thin')


def write_xlsx(file, data_list, fieldnames):
    # Specify XLSX file path
    xlsx_file_path = file

    workbook = Workbook()

    # Get the active worksheet
    sheet = workbook.active

    # Write column names

    sheet.append(fieldnames)

    format_res = []
    for i in data_list:
        row = []
        for field in fieldnames:
            # print(field,getattr(i,field))
            row.append(getattr(i, field))
        format_res.append(row)

    # Write data from PromptQA objects to the worksheet
    for row in format_res:
        sheet.append(row)

    # Save the workbook to XLSX file
    workbook.save(xlsx_file_path)

    print("Data written to", xlsx_file_path)


def write_csv(file, data_list, fieldnames, encodeing="gbk"):
    # 写入CSV文件
    csv_file_path = file
    with open(csv_file_path, mode="w", newline="", encoding=encodeing) as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

        writer.writeheader()
        for instance in data_list:
            writer.writerow(instance.__dict__)

    print(f"CSV文件已创建：{csv_file_path}")


def DateStr():
    # 获取当前日期和时间
    now = datetime.now()

    # 提取年、月、日、小时和分钟
    year = now.year
    month = now.month
    day = now.day
    hour = now.hour
    minute = now.minute

    # 格式化为字符串
    date_string = f"{year}{month:02d}{day:02d}"
    time_string = f"{hour:02d}{minute:02d}"

    # 合并日期和时间字符串
    full_string = f"{date_string}{time_string}"

    # print("完整字符串:", full_string)
    return full_string


class FusionJson():
    def compute_similiar(self, qas):
        # 中文句向量模型(CoSENT)，中文语义匹配任务推荐，支持fine-tune继续训练
        # sim_model = Similarity("shibing624/text2vec-base-chinese", SimilarityType.COSINE)
        # sim_model = Similarity("shibing624/text2vec-base-multilingual",SimilarityType.COSINE)

        n = len(qas)
        result = np.zeros((n, n))
        for i in range(len(qas)):

            if i % (n // 10 if n // 10 != 0 else 1) == 0:
                print("total: %d,current progress %f %s" % (n, i / n * 100, "%"))
            for j in range(len(qas)):
                if j < i:
                    score = self.sim_model.get_score(qas[i].prompt, qas[j].prompt)
                    dec = decimal.Decimal(score)
                    score = dec.quantize(decimal.Decimal('0.00'), rounding=decimal.ROUND_DOWN)
                    result[i][j] = score
                    if score > self.target:
                        print("q1: %s ,q2: %s,Score: %3.2f" % (qas[i].prompt, qas[j].prompt, score))
        return result

    def write_json(self, qas):
        np.savetxt(self.num_out_file, qas)

    def read_json_and_write(self):
        df = pd.read_json(self.src_rebuild_abspath)
        with open(self.questions_file, "w") as out_file:
            for index, row in df.iterrows():
                q = row.get("prompt")
                self.questions_data.append([index,q])
                out_file.write("%d  %s\n" % (index, q))

    def printN(self, sep, v):
        fmt = "%s"
        if v == 0:
            fmt = fmt + "%4d"
        else:
            fmt = fmt + "%3.2f"
        s = fmt % (sep, v)
        print(s, end='')

    def write_xlsx(self):
        p = pd.DataFrame(self.data)
        writer = pd.ExcelWriter(self.out_excel_file_abs)
        p.to_excel(writer, "分数", float_format="%.2f", header=True, index=True)
        writer.close()

        wb = load_workbook(self.out_excel_file_abs)
        ws = wb['分数']
        # 定义填充颜色
        fill_greater_than_095 = PatternFill(start_color='FF6100', end_color='FF6100', fill_type='solid')
        fill_between_090_and_095 = PatternFill(start_color='FF8000', end_color='FF8000', fill_type='solid')

        # 添加条件格式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                if 0.9 < cell.value < 0.95:
                    cell.fill = fill_between_090_and_095
                elif cell.value > 0.95:
                    cell.fill = fill_greater_than_095
        ws.freeze_panes = 'B2'

        qsheet = wb.create_sheet('问题', 0)
        for i in self.questions_data:
            qsheet.append(i)


        wb.save(self.out_excel_file_abs)
        wb.close()

    def __init__(self, **kwargs):
        self.questions_data = []
        self.file_date = DateStr()
        self.src_dir = kwargs["src_dir"]
        self.file_prefix = kwargs["file_prefix"]
        self.run_type = kwargs["run_type"]
        self.target = kwargs["target"]
        self.src_json_file = "%s.json" % (self.file_prefix)
        self.out_excel_file = "%s.xlsx" % (self.file_prefix)
        self.src_json_file_abspath = os.path.join(self.src_dir, self.src_json_file)
        self.out_excel_file_abs = os.path.join(self.src_dir, self.out_excel_file)
        self.src_rebuild_abspath = self.src_json_file_abspath + "_%s.%s_rebuild.json" % (self.file_date, self.run_type)

        # 中文句向量模型(CoSENT)，中文语义匹配任务推荐，支持fine-tune继续训练
        self.sim_model = Similarity("shibing624/text2vec-base-chinese", SimilarityType.COSINE)
        self.num_out_file = self.src_json_file_abspath + "_%s.%s_score.txt" % (self.file_date, self.run_type)
        self.num_modify_file = self.src_json_file_abspath + "_%s.%s_modified_score.txt" % (
            self.file_date, self.run_type)

        self.questions_file = self.src_json_file_abspath + "_%s.%s_prompt_output.txt" % (self.file_date, self.run_type)

    def num_format(self):
        data = np.loadtxt(self.num_out_file)
        self.data = data
        # 设置打印选项，保留两位小数，不进行四舍五入
        # np.savetxt(outputText, data, fmt='%.2f')

        # print(data.shape)
        width, height = range(data.shape[0]), range(data.shape[1])
        for i in width:
            # print()
            for j in height:
                if j == i:
                    data[i][j] = 1.00
                    # data[i][j] = "-"
                # if j == 0:
                #     printN("", data[i][j])
                # else:
                #     printN(" ", data[i][j])
        # Modify the data
        for i in range(data.shape[0]):
            for j in range(data.shape[1]):
                if j == i:
                    data[i][j] = 1.00

        # Define the output file path
        output_file_path = self.num_modify_file

        # Open the output file for writing
        with open(output_file_path, "w") as output_file:
            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    if j == 0:
                        output_file.write("{:.2f}".format(data[i][j]))
                    else:
                        output_file.write(" {:.2f}".format(data[i][j]))
                output_file.write("\n")

    def read_json(self):
        qas = []
        df = pd.read_json(self.src_rebuild_abspath)
        for index, row in df.iterrows():
            q = row.get("prompt")
            # a = row.get("chosen")
            # qas.append(QA(q, a))
            qas.append(q)
        return qas

    def check_json_contains_keys(self, keys_to_check):
        try:
            with open(self.src_json_file_abspath, "r") as out_file:
                j = json.load(out_file)
                for d in j[0:1]:
                    if d.get(keys_to_check[0]) and d.get(keys_to_check[1]):
                        return True
        except FileNotFoundError:
            print(f"File '{self.src_json_file_abspath}' not found.")
            return False
        except json.JSONDecodeError:
            print(f"Error decoding JSON in file '{self.src_json_file_abspath}'.")
            return False
        return False

    def rebuild(self):
        keys_to_check = ['prompt', 'chosen']

        contains_keys = self.check_json_contains_keys(keys_to_check)

        if contains_keys:
            print("skip rebuild")
            import shutil
            shutil.copyfile(self.src_json_file_abspath, self.src_rebuild_abspath)
            return

        # 从原始文件读取数据
        with open(self.src_json_file_abspath, "r", encoding="utf-8") as original_file:
            original_data = json.load(original_file)

        # 构建目标数据
        target_data = []
        for data in original_data:
            target = {
                "prompt": data["question"],
                "chosen": data["answer"],
                "response": "",
                "rejected": ""
            }
            target_data.append(target)

        # 将目标数据写入文件
        with open(self.src_rebuild_abspath, "w", encoding="utf-8") as target_file:
            json.dump(target_data, target_file, ensure_ascii=False, indent=2)

        print("Conversion completed.")


class QA:
    def __init__(self, q, a):
        self.prompt = q
        self.chosen = a

    def to_json(self):
        return {"instruction": self.prompt, "output": self.chosen}
